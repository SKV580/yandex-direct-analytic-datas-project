#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import math
import datetime
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ==================== НАСТРОЙКИ ====================
# Заголовок для отчета (выводится в ячейке A1 на каждой вкладке)
REPORT_TITLE = "Отчет по проекту ОПОРА за период 1 - 31 марта 2026г."
# Название сайта для поиска файлов (используется в именах файлов)
SITE_NAME = "opora"
# Папка с исходными данными
DATA_FOLDER = "data"
# Папка для результатов
OUTPUT_FOLDER = "out"
# Минимальный расход для попадания в рекомендации
MIN_SPEND = 350.0
# Минимум кликов для рекомендаций
MIN_CLICKS = 10
# ===================================================


# Словарь переименования столбцов для отображения в Excel
COLUMN_DISPLAY_NAMES = {
    "campaign_type":    "Тип кампании",
    "campaign":         "Название кампании",
    "spend":            "Расходы, руб с НДС",
    "clicks":           "Клики",
    "conversions":      "Конверсии",
    "impressions":      "Показы",
    "bounce_rate":      "Отказы, %",
    "depth":            "Глубина просмотров",
    "cpc":              "Цена клика, CPC",
    "cpa":              "Цена конверсии, CPA",
    "cr":               "Процент конверсии, CR",
    "ctr":              "Кликабельность, CTR",
    "status":           "Статус",
    "recommendations":  "Рекомендации",
    "device":           "Тип устройства",
    "gender_age":       "Пол и возраст",
    "targeting_type":   "Условие показа",
    "placement":        "Площадка",
    "ad_group":         "Группа объявлений",
    "region":           "Регион",
    "phrase":           "Ключевая фраза",
    "can_block":        "Можно запретить",
    "title":            "Заголовок",
    "image":            "Изображение",
}

# Форматы чисел для столбцов (по display-именам)
COLUMN_NUMBER_FORMATS = {
    "Расходы, руб с НДС":       '#,##0 [$₽-419]',
    "Цена клика, CPC":          '#,##0.0 [$₽-419]',
    "Цена конверсии, CPA":      '#,##0 [$₽-419]',
    "Клики":                    '#,##0',
    "Конверсии":                '#,##0',
    "Показы":                   '#,##0',
    "Отказы, %":                '0.0%',
    "Глубина просмотров":       '0.0',
    "Процент конверсии, CR":    '0.0%',
    "Кликабельность, CTR":      '0.0%',
}


CANONICAL_COLUMNS = {
    "campaign": [
        "Название кампании", "Кампания", "campaign", "campaign name"
    ],
    "ad_group": [
        "Название группы", "Группа", "Группа объявлений", "ad group"
    ],
    "placement": [
        "Название площадки",
        "Название площадок yandex",
        "Название площадки yandex",
        "Площадка",
        "Site",
        "Placement"
    ],
    "can_block": [
        "Можно запретить площадку", "can block", "blockable"
    ],
    "device": [
        "Тип устройства", "Устройство", "device"
    ],
    "gender": [
        "Пол", "gender"
    ],
    "age": [
        "Возраст", "age"
    ],
    "targeting_type": [
        "Тип условия показа", "Условие показа", "targeting type"
    ],
    "title": [
        "Заголовок", "Заголовок 1", "title"
    ],
    "image": [
        "Изображение",
        "image",
        "Название файла изображения",
        "Название изображения",
    ],
    "spend": [
        "Расход, ₽", "Расход", "Cost", "Spend"
    ],
    "clicks": [
        "Клики", "Clicks"
    ],
    "conversions": [
        "Конверсии", "Conversions"
    ],
    "impressions": [
        "Показы", "Impressions"
    ],
    "bounce_rate": [
        "Отказы, %", "Показатель отказов, %", "Отказы", "Bounce rate"
    ],
    "depth": [
        "Глубина просмотра", "Глубина просмотров", "Depth"
    ],
    "region": [
        "Регион местонахождения", "Регион", "Region"
    ],
    "phrase": [
        "Ключевая фраза", "Фраза", "Phrase", "Keyword"
    ],
}

DIMENSIONS = []


class AnalyzerError(Exception):
    pass


def normalize_text(s: str) -> str:
    s = str(s).strip().lower()
    s = s.replace("ё", "е")
    s = re.sub(r"\s+", " ", s)
    return s


def find_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    normalized_map = {normalize_text(c): c for c in df.columns}
    for c in candidates:
        if normalize_text(c) in normalized_map:
            return normalized_map[normalize_text(c)]
    return None


def clean_numeric(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.replace("\u00a0", " ", regex=False)
    s = s.str.replace("₽", "", regex=False)
    s = s.str.replace("%", "", regex=False)
    s = s.str.replace(" ", "", regex=False)
    s = s.str.replace(",", ".", regex=False)
    s = s.replace({"nan": np.nan, "None": np.nan, "": np.nan, "-": np.nan})
    return pd.to_numeric(s, errors="coerce")


def find_site_files(site_name: str, data_folder: str) -> Dict[str, Optional[Path]]:
    """Находит файлы для указанного сайта по паттернам."""
    folder = Path(data_folder)
    if not folder.exists():
        raise AnalyzerError(f"Папка {folder} не найдена")

    file_patterns = {
        "mainstat":   f"mainstat_{site_name}",
        "regions":    f"regions_{site_name}",
        "polvozrast": f"polvozrast_{site_name}",
        "pfrases":    f"pfrases_{site_name}",
    }

    found_files = {}
    for key, pattern in file_patterns.items():
        files = sorted([f for f in folder.iterdir()
                        if f.suffix.lower() in {".xlsx", ".xls"}
                        and f.name.lower().startswith(pattern.lower())
                        and not f.name.startswith("~$")])
        found_files[key] = files[0] if files else None

    if not found_files["mainstat"]:
        raise AnalyzerError(
            f"Обязательный файл mainstat_{site_name}*.xlsx не найден в папке {folder}"
        )

    print("Найдены файлы:")
    for key, path in found_files.items():
        if path:
            print(f"  {key}: {path.name}")
        else:
            print(f"  {key}: НЕ НАЙДЕН (вкладка не будет создана)")

    return found_files


def find_header_row(df: pd.DataFrame, required_columns: List[str]) -> int:
    """Находит строку с заголовками таблицы."""
    for idx in range(min(20, len(df))):
        row_values = df.iloc[idx].astype(str).str.lower().tolist()
        matches = sum(
            1 for col in required_columns
            if any(col.lower() in str(val).lower() for val in row_values)
        )
        if matches >= len(required_columns) * 0.6:
            return idx
    return 0


def load_data_file(path: Path, required_cols: List[str]) -> pd.DataFrame:
    """Загружает данные из файла с автоматическим пропуском служебных строк."""
    try:
        df_raw = pd.read_excel(path, header=None, sheet_name=0)
        header_row = find_header_row(df_raw, required_cols)
        df = pd.read_excel(path, header=header_row, sheet_name=0)
        df = df.dropna(how="all")
        print(f"  Загружено строк: {len(df)}, заголовки найдены в строке {header_row + 1}")
        return df
    except Exception as e:
        raise AnalyzerError(f"Не удалось загрузить файл {path.name}: {e}")


def map_columns(df: pd.DataFrame, required: List[str] = None) -> Dict[str, str]:
    """Маппинг колонок DataFrame через CANONICAL_COLUMNS."""
    result = {}
    for key, candidates in CANONICAL_COLUMNS.items():
        col = find_column(df, candidates)
        if col:
            result[key] = col

    if required is None:
        required = ["campaign", "spend", "clicks", "conversions"]

    missing = [k for k in required if k not in result]
    if missing:
        raise AnalyzerError(f"Не найдены обязательные колонки: {', '.join(missing)}")
    return result


def prepare_dataframe(df: pd.DataFrame, mapping: Dict[str, str]) -> pd.DataFrame:
    work = pd.DataFrame()
    for key, col in mapping.items():
        work[key] = df[col]

    if "campaign" in work.columns:
        work = work[work["campaign"].notna()]
        work = work[work["campaign"].astype(str).str.strip() != ""]
        work = work[~work["campaign"].astype(str).str.lower().str.contains("итого", na=False)]

    for col in ["spend", "clicks", "conversions", "impressions", "bounce_rate", "depth"]:
        if col in work.columns:
            work[col] = clean_numeric(work[col])
        else:
            work[col] = np.nan

    if "bounce_rate" in work.columns:
        s = work["bounce_rate"].dropna()
        if not s.empty and s.max() > 1.0:
            work["bounce_rate"] = work["bounce_rate"] / 100.0

    for col in ["campaign", "ad_group", "placement", "device", "gender", "age",
                "targeting_type", "title", "image", "can_block", "region", "phrase"]:
        if col not in work.columns:
            work[col] = "Не указано"
        work[col] = work[col].fillna("Не указано").astype(str).str.strip().replace({"": "Не указано"})

    work["spend"] = work["spend"].fillna(0.0)
    work["clicks"] = work["clicks"].fillna(0.0)
    work["conversions"] = work["conversions"].fillna(0.0)
    work["impressions"] = work["impressions"].fillna(0.0)

    return work


def load_and_prepare_extra(path: Path, required_keys: List[str]) -> pd.DataFrame:
    """
    Универсальная загрузка дополнительного файла через CANONICAL_COLUMNS.
    required_keys — список ключей из CANONICAL_COLUMNS, которые обязательны.
    """
    # Получаем примеры названий колонок для поиска заголовка
    hint_cols = []
    for key in required_keys:
        if key in CANONICAL_COLUMNS:
            hint_cols.extend(CANONICAL_COLUMNS[key][:2])

    df_raw = load_data_file(path, hint_cols)
    mapping = map_columns(df_raw, required=required_keys)
    df = prepare_dataframe(df_raw, mapping)
    df = add_metrics(df)
    return df


def determine_campaign_type(campaign_name: str) -> str:
    """Определяет тип кампании по её названию."""
    name = str(campaign_name)
    name_lower = name.lower()

    if "динамическ" in name_lower:
        return "Динамическая"
    if "МК " in name:
        return "Мастер кампаний"
    if "смарт" in name_lower:
        return "Смарт-баннеры"
    if "товарн" in name_lower:
        return "Товарная"
    if "рся" in name_lower:
        return "РСЯ"
    if "поиск" in name_lower:
        return "Поиск"

    return "Не определен"


def add_metrics(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in ["spend", "clicks", "conversions", "impressions"]:
        if col not in df.columns:
            df[col] = 0.0

    df["cpc"] = np.where(df["clicks"] > 0, df["spend"] / df["clicks"], np.nan)
    df["cpa"] = np.where(df["conversions"] > 0, df["spend"] / df["conversions"], np.nan)
    df["cr"]  = np.where(df["clicks"] > 0, df["conversions"] / df["clicks"], np.nan)
    df["ctr"] = np.where(df["impressions"] > 0, df["clicks"] / df["impressions"], np.nan)

    if "campaign" in df.columns:
        df["campaign_type"] = df["campaign"].apply(determine_campaign_type)

    return df


def summarize(df: pd.DataFrame, dim: str, by_campaign: bool = False) -> pd.DataFrame:
    """Агрегация данных по измерению, опционально с группировкой по кампаниям."""
    group_cols = ["campaign", dim] if by_campaign else [dim]

    agg_dict = {
        "spend":       ("spend", "sum"),
        "clicks":      ("clicks", "sum"),
        "conversions": ("conversions", "sum"),
        "impressions": ("impressions", "sum"),
    }
    if "bounce_rate" in df.columns:
        agg_dict["bounce_rate"] = ("bounce_rate", "mean")
    if "depth" in df.columns:
        agg_dict["depth"] = ("depth", "mean")
    if dim == "placement" and "can_block" in df.columns:
        agg_dict["can_block"] = ("can_block", "first")

    agg = df.groupby(group_cols, dropna=False).agg(**agg_dict).reset_index()

    agg["cpc"] = np.where(agg["clicks"] > 0, agg["spend"] / agg["clicks"], np.nan)
    agg["cpa"] = np.where(agg["conversions"] > 0, agg["spend"] / agg["conversions"], np.nan)
    agg["cr"]  = np.where(agg["clicks"] > 0, agg["conversions"] / agg["clicks"], np.nan)
    agg["ctr"] = np.where(agg["impressions"] > 0, agg["clicks"] / agg["impressions"], np.nan)

    sort_cols = ["campaign", "spend"] if by_campaign else ["conversions", "spend"]
    agg = agg.sort_values(
        sort_cols,
        ascending=[True, False] if by_campaign else [False, False]
    ).reset_index(drop=True)

    return agg


def summarize_gender_age(df: pd.DataFrame, by_campaign: bool = True) -> pd.DataFrame:
    """Агрегация по полу и возрасту вместе."""
    if "gender" not in df.columns or "age" not in df.columns:
        return pd.DataFrame()

    df_copy = df.copy()
    df_copy["gender_age"] = df_copy["gender"].astype(str) + " " + df_copy["age"].astype(str)

    group_cols = ["campaign", "gender_age"] if by_campaign else ["gender_age"]

    agg_dict = {
        "spend":       ("spend", "sum"),
        "clicks":      ("clicks", "sum"),
        "conversions": ("conversions", "sum"),
        "impressions": ("impressions", "sum"),
    }
    if "bounce_rate" in df_copy.columns:
        agg_dict["bounce_rate"] = ("bounce_rate", "mean")
    if "depth" in df_copy.columns:
        agg_dict["depth"] = ("depth", "mean")

    agg = df_copy.groupby(group_cols, dropna=False).agg(**agg_dict).reset_index()

    agg["cpc"] = np.where(agg["clicks"] > 0, agg["spend"] / agg["clicks"], np.nan)
    agg["cpa"] = np.where(agg["conversions"] > 0, agg["spend"] / agg["conversions"], np.nan)
    agg["cr"]  = np.where(agg["clicks"] > 0, agg["conversions"] / agg["clicks"], np.nan)
    agg["ctr"] = np.where(agg["impressions"] > 0, agg["clicks"] / agg["impressions"], np.nan)

    if by_campaign:
        agg = agg.sort_values(["campaign", "spend"], ascending=[True, False]).reset_index(drop=True)
    else:
        agg = agg.sort_values("spend", ascending=False).reset_index(drop=True)

    return agg


def overall_metrics(df: pd.DataFrame) -> Dict[str, float]:
    spend  = df["spend"].sum()
    clicks = df["clicks"].sum()
    conv   = df["conversions"].sum()
    impr   = df["impressions"].sum()
    return {
        "spend":       spend,
        "clicks":      clicks,
        "conversions": conv,
        "impressions": impr,
        "cpc":  spend / clicks if clicks else np.nan,
        "cpa":  spend / conv   if conv   else np.nan,
        "cr":   conv  / clicks if clicks else np.nan,
        "ctr":  clicks / impr  if impr   else np.nan,
    }


def best_and_worst(
    summary: pd.DataFrame, min_clicks: int, min_spend: float
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    eligible = summary[
        (summary["clicks"] >= min_clicks) | (summary["spend"] >= min_spend)
    ].copy()
    if eligible.empty:
        return eligible, eligible
    good = (
        eligible[eligible["conversions"] > 0]
        .sort_values(["cpa", "conversions"], ascending=[True, False])
        .head(5)
    )
    bad = (
        eligible[
            (eligible["conversions"] == 0)
            | (eligible["cpa"] > eligible["cpa"].median())
        ]
        .sort_values(["conversions", "spend"], ascending=[True, False])
        .head(5)
    )
    return good, bad


def campaign_recommendations(
    df: pd.DataFrame,
    campaign_summary: pd.DataFrame,
    min_clicks: int,
    min_spend: float,
    available_dims: set,
    additional_slices: dict = None,
) -> pd.DataFrame:
    """Генерация рекомендаций по кампаниям."""
    account_cpa = overall_metrics(df)["cpa"]
    rows = []

    if additional_slices is None:
        additional_slices = {}

    for _, row in campaign_summary.iterrows():
        camp = row["campaign"]
        part = df[df["campaign"] == camp].copy()
        recs = []

        for dim, label in [
            ("device",         "устройствам"),
            ("gender",         "полу"),
            ("age",            "возрасту"),
            ("targeting_type", "типу условия показа"),
            ("placement",      "площадкам"),
        ]:
            if dim not in available_dims:
                continue
            s = summarize(part, dim)
            good, bad = best_and_worst(s, min_clicks=min_clicks, min_spend=min_spend)
            if not good.empty and not bad.empty:
                bad_set = set(bad[dim].astype(str).tolist())
                good = good[~good[dim].astype(str).isin(bad_set)]
            if not good.empty:
                recs.append(f"Усилить по {label}: {', '.join(good[dim].astype(str).head(3).tolist())}")
            if not bad.empty:
                recs.append(f"Сократить по {label}: {', '.join(bad[dim].astype(str).head(3).tolist())}")

        for slice_key, dim_col, label in [
            ("regions",    "region",     "регионам"),
            ("gender_age", "gender_age", "полу и возрасту"),
            ("phrases",    "phrase",     "ключевым фразам"),
        ]:
            if slice_key not in additional_slices:
                continue
            slice_data = additional_slices[slice_key]
            camp_data  = slice_data[slice_data["campaign"] == camp]
            if camp_data.empty:
                continue
            good, bad = best_and_worst(
                camp_data.rename(columns={dim_col: "_dim"}),
                min_clicks=min_clicks, min_spend=min_spend,
            )
            if not good.empty:
                good = good.rename(columns={"_dim": dim_col})
            if not bad.empty:
                bad  = bad.rename(columns={"_dim": dim_col})
            if not good.empty and not bad.empty:
                bad_set = set(bad[dim_col].astype(str).tolist())
                good = good[~good[dim_col].astype(str).isin(bad_set)]
            n = 5 if slice_key == "phrases" else 3
            if not good.empty:
                recs.append(f"Усилить по {label}: {', '.join(good[dim_col].astype(str).head(n).tolist())}")
            if not bad.empty:
                recs.append(f"Сократить по {label}: {', '.join(bad[dim_col].astype(str).head(n).tolist())}")

        status = "усилить"
        if pd.notna(row["cpa"]) and pd.notna(account_cpa):
            if row["cpa"] > account_cpa * 1.5:
                status = "сократить и пересобрать"
            elif row["cpa"] > account_cpa * 1.15:
                status = "оптимизировать"

        rows.append({
            "campaign":        camp,
            "spend":           row["spend"],
            "clicks":          row["clicks"],
            "conversions":     row["conversions"],
            "cpc":             row["cpc"],
            "cpa":             row["cpa"],
            "cr":              row["cr"],
            "status":          status,
            "recommendations": " | ".join(recs) if recs else "Недостаточно данных для рекомендаций",
        })

    return pd.DataFrame(rows).sort_values(
        ["conversions", "spend"], ascending=[False, False]
    )


def print_terminal_report(
    df: pd.DataFrame, campaign_summary: pd.DataFrame, recs: pd.DataFrame
) -> None:
    totals = overall_metrics(df)
    print("=" * 90)
    print("КРАТКИЙ ОТЧЕТ")
    print("=" * 90)
    print(f"Общий расход: {totals['spend']:.2f} руб")
    print(f"Клики: {int(totals['clicks'])}")
    print(f"Конверсии: {int(totals['conversions'])}")
    print(f"CPC: {totals['cpc']:.2f} руб")
    print(f"CPA: {totals['cpa']:.2f} руб")
    print(f"CR: {totals['cr']*100:.2f}%")
    print(f"\nВсего кампаний: {len(campaign_summary)}")
    print("Полный отчет сохранен в Excel и TXT файлах")
    print("=" * 90)


def safe_sheet_name(name: str) -> str:
    name = re.sub(r"[\\/*?:\[\]]", "_", str(name))
    return name[:31]


def rename_columns_for_display(df: pd.DataFrame) -> pd.DataFrame:
    """Переименовывает внутренние имена колонок в читаемые для Excel."""
    return df.rename(columns={k: v for k, v in COLUMN_DISPLAY_NAMES.items() if k in df.columns})


def write_df_sheet(
    ws, df: pd.DataFrame, table_name: str, sheet_title: str, report_title: str
) -> None:
    """Записывает DataFrame на лист Excel с заголовками."""
    ws["A1"] = report_title
    ws["A1"].font = Font(size=18, bold=True)

    ws["A2"] = sheet_title
    ws["A2"].font = Font(size=16, bold=True)

    if df.empty:
        ws["A3"] = "Нет данных"
        return

    # Переименовываем колонки для отображения
    df = rename_columns_for_display(df.copy())

    start_row, start_col = 4, 1
    headers = list(df.columns)

    for i, h in enumerate(headers, start=start_col):
        c = ws.cell(row=start_row, column=i, value=h)
        c.font      = Font(bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, val in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=None if (isinstance(val, float) and math.isnan(val)) else val)

    end_row = start_row + len(df)
    end_col = start_col + len(headers) - 1
    ref = (
        f"{get_column_letter(start_col)}{start_row}:"
        f"{get_column_letter(end_col)}{end_row}"
    )
    tab = Table(displayName=table_name[:25], ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tab)
    ws.freeze_panes = f"{get_column_letter(start_col)}{start_row + 1}"

    # Применяем форматы и ширины колонок
    for idx, h in enumerate(headers, start=start_col):
        col_letter = get_column_letter(idx)
        max_len = (
            df.iloc[:, idx - start_col]
            .apply(lambda x: len(str(x)) if pd.notna(x) else 0)
            .max()
            if len(df) else 0
        )
        width = max(12, min(45, max(len(str(h)), max_len) + 2))
        ws.column_dimensions[col_letter].width = width

        # Числовые форматы из словаря
        fmt = COLUMN_NUMBER_FORMATS.get(h)
        if fmt:
            for cell in ws[col_letter][start_row:end_row]:
                cell.number_format = fmt

    # Условное форматирование — цветовые шкалы
    spend_col = "Расходы, руб с НДС"
    cpa_col   = "Цена конверсии, CPA"

    if spend_col in headers:
        col_letter = get_column_letter(start_col + headers.index(spend_col))
        ws.conditional_formatting.add(
            f"{col_letter}{start_row+1}:{col_letter}{end_row}",
            ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                end_type="max",   end_color="BDD7EE",
            ),
        )
    if cpa_col in headers:
        col_letter = get_column_letter(start_col + headers.index(cpa_col))
        ws.conditional_formatting.add(
            f"{col_letter}{start_row+1}:{col_letter}{end_row}",
            ColorScaleRule(
                start_type="min",        start_color="63BE7B",
                mid_type="percentile",   mid_value=50, mid_color="FFEB84",
                end_type="max",          end_color="F8696B",
            ),
        )


def _overview_thin_border() -> Border:
    thin = Side(style="thin", color="D9D9D9")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def write_overview_sheet(ws, totals: Dict[str, float], report_title: str) -> None:
    """Красивое оформление вкладки Обзор."""
    # ── Заголовки ──────────────────────────────────────────────────────────────
    ws["A1"] = report_title
    ws["A1"].font = Font(size=18, bold=True, color="1F4E78")

    ws["A2"] = "Обзор"
    ws["A2"].font = Font(size=16, bold=True, color="2E75B6")

    # Цвета карточек (чередуются)
    CARD_FILLS = [
        PatternFill("solid", fgColor="1F4E78"),  # тёмно-синий
        PatternFill("solid", fgColor="2E75B6"),  # синий
        PatternFill("solid", fgColor="2E75B6"),
        PatternFill("solid", fgColor="1F4E78"),
        PatternFill("solid", fgColor="2E75B6"),
        PatternFill("solid", fgColor="1F4E78"),
        PatternFill("solid", fgColor="2E75B6"),
        PatternFill("solid", fgColor="1F4E78"),
    ]

    metrics = [
        ("Общий расход, руб с НДС", totals["spend"],       '#,##0 [$₽-419]'),
        ("Показы",                  totals["impressions"],  '#,##0'),
        ("Клики",                   totals["clicks"],       '#,##0'),
        ("Конверсии",               totals["conversions"],  '#,##0'),
        ("Средний CPC",             totals["cpc"],          '#,##0.0 [$₽-419]'),
        ("Средний CPA",             totals["cpa"],          '#,##0 [$₽-419]'),
        ("CR (конверсия)",          totals["cr"],           '0.0%'),
        ("CTR (кликабельность)",    totals["ctr"],          '0.0%'),
    ]

    # Размещаем по 4 карточки в ряд, 2 ряда
    # Каждая карточка: 2 строки высотой, 2 колонки шириной
    # Стартовая строка карточек — 4
    card_start_row = 4
    cards_per_row  = 4
    col_width      = 22   # ширина каждой из 8 колонок-карточек

    border = _overview_thin_border()

    for i, (label, value, fmt) in enumerate(metrics):
        card_col = (i % cards_per_row) * 2 + 1          # 1,3,5,7 / 1,3,5,7
        card_row = card_start_row + (i // cards_per_row) * 3  # 4 / 7

        fill = CARD_FILLS[i]

        # Строка 1: метка
        lbl_cell = ws.cell(row=card_row, column=card_col, value=label)
        lbl_cell.font      = Font(bold=True, color="FFFFFF", size=10)
        lbl_cell.fill      = fill
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        lbl_cell.border    = border

        # Строка 2: значение
        val_cell = ws.cell(row=card_row + 1, column=card_col,
                           value=None if pd.isna(value) else float(value))
        val_cell.font          = Font(bold=True, color="FFFFFF", size=16)
        val_cell.fill          = fill
        val_cell.alignment     = Alignment(horizontal="center", vertical="center")
        val_cell.number_format = fmt
        val_cell.border        = border

        # Строка 3: пустая разделительная строка (светло-серая)
        sep_cell = ws.cell(row=card_row + 2, column=card_col)
        sep_cell.fill   = PatternFill("solid", fgColor="F2F2F2")
        sep_cell.border = border

        # Объединяем ячейки по 2 столбца для каждой карточки
        for r_offset in range(3):
            ws.merge_cells(
                start_row=card_row + r_offset, start_column=card_col,
                end_row=card_row + r_offset,   end_column=card_col + 1,
            )

        # Высота строк карточки
        ws.row_dimensions[card_row].height     = 20
        ws.row_dimensions[card_row + 1].height = 36
        ws.row_dimensions[card_row + 2].height = 8

    # Ширины колонок (8 колонок по 2 → колонки A..P)
    for col_i in range(1, cards_per_row * 2 + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = col_width // 2 + 2

    # Небольшой отступ снизу после карточек + подсказка
    note_row = card_start_row + 7
    ws.cell(row=note_row, column=1,
            value="Подробные данные — на следующих вкладках").font = Font(
        italic=True, color="7F7F7F", size=10
    )


def generate_recommendations_txt(
    output_path: Path,
    campaign_summary: pd.DataFrame,
    recs: pd.DataFrame,
    placements_data: pd.DataFrame,
    report_title: str,
) -> None:
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(f"{report_title}\n")
        f.write("=" * 80 + "\n\n")
        f.write("РЕКОМЕНДАЦИИ ПО УЛУЧШЕНИЮ КАМПАНИЙ\n")
        f.write("=" * 80 + "\n\n")

        for _, row in recs.iterrows():
            campaign = row["campaign"]
            f.write(f"\nКАМПАНИЯ: {campaign}\n")
            f.write("-" * 80 + "\n")
            f.write(f"Статус: {row['status']}\n")
            f.write(f"Расход: {row['spend']:.2f} руб\n")
            f.write(f"Конверсии: {int(row['conversions'])}\n")
            cpa_val = row['cpa']
            f.write(f"CPA: {cpa_val:.2f} руб\n\n" if pd.notna(cpa_val) else "CPA: —\n\n")

            f.write("Рекомендации:\n")
            for rec in str(row["recommendations"]).split(" | "):
                f.write(f"  • {rec}\n")

            if not placements_data.empty:
                campaign_placements = placements_data[placements_data["campaign"] == campaign]
                bad_placements = campaign_placements[
                    (campaign_placements["conversions"] == 0)
                    & (campaign_placements["spend"] > 100)
                ].copy()

                if not bad_placements.empty:
                    f.write("\nПЛОЩАДКИ ДЛЯ МИНУСОВКИ:\n")
                    for _, pl_row in bad_placements.iterrows():
                        placement_name = pl_row.get("placement", "Неизвестная площадка")
                        can_block      = pl_row.get("can_block", "Да")
                        spend          = pl_row.get("spend", 0)
                        if str(can_block).lower() == "нет":
                            f.write(
                                f"  • {placement_name} (расход: {spend:.2f} руб)"
                                f" - ЭТУ ПЛОЩАДКУ ВЫКЛЮЧИТЬ НЕЛЬЗЯ\n"
                            )
                        else:
                            f.write(f"  • {placement_name} (расход: {spend:.2f} руб)\n")
            f.write("\n")

        f.write("\n" + "=" * 80 + "\n")
        f.write("Конец отчета\n")


def build_excel(
    output_path: Path,
    totals: Dict[str, float],
    campaign_summary: pd.DataFrame,
    recs: pd.DataFrame,
    slices: Dict[str, pd.DataFrame],
    report_title: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Обзор"

    # ── Обзор (красивые карточки) ───────────────────────────────────────────
    write_overview_sheet(ws, totals, report_title)

    # ── Кампании ─────────────────────────────────────────────────────────────
    camp_sheet = wb.create_sheet("Кампании")
    write_df_sheet(camp_sheet, campaign_summary, "CampaignSummary", "Кампании", report_title)

    # ── Тип кампании (вторым после Обзора) ───────────────────────────────────
    if "campaign_type" in slices:
        type_sheet = wb.create_sheet("Тип кампании")
        write_df_sheet(type_sheet, slices["campaign_type"], "CampaignType", "Тип кампании", report_title)

    # ── Рекомендации ─────────────────────────────────────────────────────────
    rec_sheet = wb.create_sheet("Рекомендации")
    write_df_sheet(rec_sheet, recs, "CampaignRecs", "Рекомендации", report_title)

    # ── Остальные вкладки ─────────────────────────────────────────────────────
    # Порядок: Устройство → Устройства-все-камп → Пол и возраст → Пол и возраст-все-камп → …
    ordered_keys = [
        ("device",          "Устройство",              "Устройство (по кампаниям)"),
        ("device_all",      "Устройства-все-камп",     "Устройства по всем кампаниям"),
        ("gender_age",      "Пол и возраст",           "Пол и возраст (по кампаниям)"),
        ("gender_age_all",  "Пол и возраст-все-камп",  "Пол и возраст по всем кампаниям"),
        ("targeting_type",  "Условие показа",          "Условие показа (по кампаниям)"),
        ("placement",       "Площадка",                "Площадка (по кампаниям)"),
        ("ad_group",        "Группа объявлений",       "Группа объявлений (по кампаниям)"),
        ("title",           "Заголовок",               "Заголовок (по кампаниям)"),
        ("image",           "Изображение",             "Изображение (по кампаниям)"),
        ("regions",         "Регионы",                 "Регионы (по кампаниям)"),
        ("phrases",         "Поисковые фразы",         "Поисковые фразы (по кампаниям)"),
    ]

    for key, sheet_name, sheet_title in ordered_keys:
        if key in slices and not slices[key].empty:
            sh = wb.create_sheet(safe_sheet_name(sheet_name))
            write_df_sheet(sh, slices[key], f"T_{key[:20]}", sheet_title, report_title)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    print("=" * 80)
    print("АНАЛИЗАТОР ОТЧЕТОВ ЯНДЕКС.ДИРЕКТ")
    print("=" * 80)
    print(f"\nНастройки:")
    print(f"  Сайт: {SITE_NAME}")
    print(f"  Заголовок отчета: {REPORT_TITLE}")
    print(f"  Папка с данными: {DATA_FOLDER}")
    print(f"  Папка для результатов: {OUTPUT_FOLDER}\n")

    try:
        # ── 1. Поиск файлов ──────────────────────────────────────────────────
        print("Шаг 1: Поиск файлов данных...")
        files = find_site_files(SITE_NAME, DATA_FOLDER)
        print()

        # ── 2. Основной файл ─────────────────────────────────────────────────
        print("Шаг 2: Загрузка основного файла mainstat...")
        df_main   = load_data_file(files["mainstat"], ["Название кампании", "Расход", "Клики", "Конверсии"])
        mapping   = map_columns(df_main)
        df        = prepare_dataframe(df_main, mapping)
        df        = add_metrics(df)
        print(f"  Обработано кампаний: {df['campaign'].nunique()}")
        print()

        # ── 3. Дополнительные файлы ──────────────────────────────────────────
        additional_slices = {}

        if files["regions"]:
            print("Шаг 3a: Загрузка данных по регионам...")
            try:
                df_reg = load_and_prepare_extra(files["regions"], ["campaign", "region", "spend", "clicks", "conversions"])
                # Удаляем строки «Итого»
                df_reg = df_reg[~df_reg["campaign"].str.lower().str.contains("итого", na=False)]
                df_reg = df_reg[df_reg["campaign"].str.strip() != ""]
                additional_slices["regions"] = summarize(df_reg, "region", by_campaign=True)
                print(f"  [OK] Регионов: {df_reg['region'].nunique()}")
            except Exception as e:
                print(f"  [ОШИБКА] regions: {e}")

        if files["polvozrast"]:
            print("Шаг 3b: Загрузка данных по полу и возрасту...")
            try:
                df_pv = load_and_prepare_extra(files["polvozrast"], ["campaign", "gender", "age", "spend", "clicks", "conversions"])
                df_pv = df_pv[~df_pv["campaign"].str.lower().str.contains("итого", na=False)]
                df_pv = df_pv[df_pv["campaign"].str.strip() != ""]
                additional_slices["gender_age"]     = summarize_gender_age(df_pv, by_campaign=True)
                additional_slices["gender_age_all"] = summarize_gender_age(df_pv, by_campaign=False)
                print(f"  [OK] Комбинаций пол+возраст: {len(additional_slices['gender_age'])}")
            except Exception as e:
                print(f"  [ОШИБКА] polvozrast: {e}")

        if files["pfrases"]:
            print("Шаг 3c: Загрузка данных по поисковым фразам...")
            try:
                df_pf = load_and_prepare_extra(files["pfrases"], ["campaign", "phrase", "spend", "clicks", "conversions"])
                df_pf = df_pf[~df_pf["campaign"].str.lower().str.contains("итого", na=False)]
                df_pf = df_pf[df_pf["campaign"].str.strip() != ""]
                additional_slices["phrases"] = summarize(df_pf, "phrase", by_campaign=True)
                print(f"  [OK] Фраз: {df_pf['phrase'].nunique()}")
            except Exception as e:
                print(f"  [ОШИБКА] pfrases: {e}")
        print()

        # ── 4. Метрики ───────────────────────────────────────────────────────
        print("Шаг 4: Расчет метрик...")
        totals           = overall_metrics(df)
        campaign_summary = summarize(df, "campaign")
        available_dims   = set(mapping.keys())

        # ── 5. Срезы из основного файла ──────────────────────────────────────
        print("Шаг 5: Создание аналитических срезов...")
        slices = {}

        slices["campaign_type"] = summarize(df, "campaign_type")
        print("  [OK] Тип кампании")

        if "device" in available_dims:
            slices["device"]     = summarize(df, "device", by_campaign=True)
            slices["device_all"] = summarize(df, "device", by_campaign=False)
            print("  [OK] Устройство (по кампаниям + суммарно)")

        # Пол и возраст из основного файла (если нет отдельного)
        if "gender" in available_dims and "age" in available_dims and "gender_age" not in additional_slices:
            slices["gender_age"]     = summarize_gender_age(df, by_campaign=True)
            slices["gender_age_all"] = summarize_gender_age(df, by_campaign=False)
            print("  [OK] Пол и возраст из основного файла")

        if "targeting_type" in available_dims:
            slices["targeting_type"] = summarize(df, "targeting_type", by_campaign=True)
            print("  [OK] Условие показа (по кампаниям)")

        if "placement" in available_dims:
            slices["placement"] = summarize(df, "placement", by_campaign=True)
            print("  [OK] Площадка (по кампаниям)")

        if "ad_group" in available_dims:
            slices["ad_group"] = summarize(df, "ad_group", by_campaign=True)
            print("  [OK] Группа объявлений (по кампаниям)")

        if "title" in available_dims:
            slices["title"] = summarize(df, "title", by_campaign=True)
            print("  [OK] Заголовок (по кампаниям)")

        if "image" in available_dims:
            slices["image"] = summarize(df, "image", by_campaign=True)
            print("  [OK] Изображение (по кампаниям)")

        # Добавляем срезы из доп. файлов
        for key, data in additional_slices.items():
            if key not in slices:
                slices[key] = data
                print(f"  [OK] {key} из отдельного файла")
        print()

        # ── 6. Рекомендации ──────────────────────────────────────────────────
        print("Шаг 6: Формирование рекомендаций...")
        recs = campaign_recommendations(
            df, campaign_summary, MIN_CLICKS, MIN_SPEND,
            available_dims=available_dims,
            additional_slices=additional_slices,
        )
        print()

        # ── 7. Терминал ──────────────────────────────────────────────────────
        print_terminal_report(df, campaign_summary, recs)

        # ── 8. Excel ─────────────────────────────────────────────────────────
        print("Шаг 7: Создание Excel-отчета...")
        output_excel = Path(OUTPUT_FOLDER) / f"{SITE_NAME}_analysis_result.xlsx"
        try:
            build_excel(output_excel, totals, campaign_summary, recs, slices, REPORT_TITLE)
            print(f"  [OK] Excel-отчет сохранен: {output_excel}")
        except PermissionError:
            print(f"  [ОШИБКА] Файл {output_excel} открыт в другой программе!")
            raise

        # ── 9. TXT ───────────────────────────────────────────────────────────
        print("Шаг 8: Создание TXT-файла с рекомендациями...")
        output_txt      = Path(OUTPUT_FOLDER) / f"{SITE_NAME}_recommendations.txt"
        placements_data = slices.get("placement", pd.DataFrame())
        generate_recommendations_txt(output_txt, campaign_summary, recs, placements_data, REPORT_TITLE)
        print(f"  [OK] TXT-файл сохранен: {output_txt}")

        print("\n" + "=" * 80)
        print("АНАЛИЗ ЗАВЕРШЕН УСПЕШНО!")
        print("=" * 80)

    except AnalyzerError as e:
        print(f"\n[ERROR] ОШИБКА: {e}")
        return 1
    except PermissionError:
        print("\n[ERROR] ФАЙЛ ЗАБЛОКИРОВАН. Закройте Excel-файл и повторите попытку.")
        return 1
    except Exception as e:
        print(f"\n[ERROR] НЕОЖИДАННАЯ ОШИБКА: {e}")
        import traceback
        traceback.print_exc()
        return 1

    return 0


if __name__ == "__main__":
    main()