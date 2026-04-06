#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import math
import datetime
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ==================== НАСТРОЙКИ ====================
# Заголовок для отчета (выводится в ячейке A1 на каждой вкладке)
REPORT_TITLE = "Отчет по сайту Опора за период 1 марта - 04 апреля 2026г."
# Название сайта для поиска файлов (используется в именах файлов)
SITE_NAME = "opora"
# Папка с исходными данными
DATA_FOLDER = "data"
# Папка для результатов
OUTPUT_FOLDER = "out"
# Минимальный расход для попадания в рекомендации
MIN_SPEND = 500.0
# Минимум кликов для рекомендаций
MIN_CLICKS = 20
# ===================================================


CANONICAL_COLUMNS = {
    "campaign": [
        "Название кампании", "Кампания", "campaign", "campaign name"
    ],
    "ad_group": [
        "Название группы", "Группа", "Группа объявлений", "ad group"
    ],
    "placement": [
        "Название площадки",
        "Название площадок yandex",
        "Название площадки yandex",
        "Площадка",
        "Site",
        "Placement"
    ],
    "can_block": [
        "Можно запретить площадку", "can block", "blockable"
    ],
    "device": [
        "Тип устройства", "Устройство", "device"
    ],
    "gender": [
        "Пол", "gender"
    ],
    "age": [
        "Возраст", "age"
    ],
    "targeting_type": [
        "Тип условия показа", "Условие показа", "targeting type"
    ],
    "title": [
        "Заголовок", "Заголовок 1", "title"
    ],
    "image": [
        "Изображение",
        "image",
        "Название файла изображения"
    ],
    "spend": [
        "Расход, ₽", "Расход", "Cost", "Spend"
    ],
    "clicks": [
        "Клики", "Clicks"
    ],
    "conversions": [
        "Конверсии", "Conversions"
    ],
    "impressions": [
        "Показы", "Impressions"
    ],
    "bounce_rate": [
        "Отказы, %", "Показатель отказов, %", "Bounce rate"
    ],
    "depth": [
        "Глубина просмотра", "Depth"
    ],
}

# Этот список больше не используется напрямую, 
# так как мы создаем срезы вручную в main() с нужными параметрами
DIMENSIONS = []


class AnalyzerError(Exception):
    pass


def normalize_text(s: str) -> str:
    s = str(s).strip().lower()
    s = s.replace("ё", "е")
    s = re.sub(r"\s+", " ", s)
    return s


def find_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    normalized_map = {normalize_text(c): c for c in df.columns}
    for c in candidates:
        if normalize_text(c) in normalized_map:
            return normalized_map[normalize_text(c)]
    return None


def clean_numeric(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.replace("\u00a0", " ", regex=False)
    s = s.str.replace("₽", "", regex=False)
    s = s.str.replace("%", "", regex=False)
    s = s.str.replace(" ", "", regex=False)
    s = s.str.replace(",", ".", regex=False)
    s = s.replace({"nan": np.nan, "None": np.nan, "": np.nan, "-": np.nan})
    return pd.to_numeric(s, errors="coerce")


def find_site_files(site_name: str, data_folder: str) -> Dict[str, Optional[Path]]:
    """Находит файлы для указанного сайта по паттернам."""
    folder = Path(data_folder)
    if not folder.exists():
        raise AnalyzerError(f"Папка {folder} не найдена")
    
    file_patterns = {
        "mainstat": f"mainstat_{site_name}",
        "regions": f"regions_{site_name}",
        "polvozrast": f"polvozrast_{site_name}",
        "pfrases": f"pfrases_{site_name}",
    }
    
    found_files = {}
    for key, pattern in file_patterns.items():
        files = sorted([f for f in folder.iterdir() 
                       if f.suffix.lower() in {".xlsx", ".xls"} 
                       and f.name.lower().startswith(pattern.lower())
                       and not f.name.startswith("~$")])
        found_files[key] = files[0] if files else None
    
    # Проверяем обязательный файл mainstat
    if not found_files["mainstat"]:
        raise AnalyzerError(f"Обязательный файл mainstat_{site_name}*.xlsx не найден в папке {folder}")
    
    print(f"Найдены файлы:")
    for key, path in found_files.items():
        if path:
            print(f"  {key}: {path.name}")
        else:
            print(f"  {key}: НЕ НАЙДЕН (вкладка не будет создана)")
    
    return found_files


def find_header_row(df: pd.DataFrame, required_columns: List[str]) -> int:
    """Находит строку с заголовками таблицы."""
    for idx in range(min(20, len(df))):
        row_values = df.iloc[idx].astype(str).str.lower().tolist()
        matches = sum(1 for col in required_columns if any(col.lower() in str(val).lower() for val in row_values))
        if matches >= len(required_columns) * 0.6:  # Если нашли хотя бы 60% требуемых колонок
            return idx
    return 0

def load_data_file(path: Path, required_cols: List[str]) -> pd.DataFrame:
    """Загружает данные из файла с автоматическим пропуском служебных строк."""
    try:
        # Сначала читаем без заголовков, чтобы найти строку с заголовками
        df_raw = pd.read_excel(path, header=None, sheet_name=0)
        header_row = find_header_row(df_raw, required_cols)
        
        # Теперь читаем с правильной строки заголовков
        df = pd.read_excel(path, header=header_row, sheet_name=0)
        
        # Удаляем полностью пустые строки
        df = df.dropna(how='all')
        
        print(f"  Загружено строк: {len(df)}, заголовки найдены в строке {header_row + 1}")
        return df
    except Exception as e:
        raise AnalyzerError(f"Не удалось загрузить файл {path.name}: {e}")


def map_columns(df: pd.DataFrame) -> Dict[str, str]:
    result = {}
    for key, candidates in CANONICAL_COLUMNS.items():
        col = find_column(df, candidates)
        if col:
            result[key] = col
    required = ["campaign", "spend", "clicks", "conversions"]
    missing = [k for k in required if k not in result]
    if missing:
        raise AnalyzerError(f"Не найдены обязательные колонки: {', '.join(missing)}")
    return result


def prepare_dataframe(df: pd.DataFrame, mapping: Dict[str, str]) -> pd.DataFrame:
    work = pd.DataFrame()
    for key, col in mapping.items():
        work[key] = df[col]

    # Удаляем строки "Итого" - они определяются по пустым значениям в критических полонках
    # Строка Итого обычно имеет пустые "Название кампании" или агрегирует все данные
    if "campaign" in work.columns:
        # Удаляем строки где кампания пустая или равна "Итого"
        work = work[work["campaign"].notna()]
        work = work[work["campaign"].astype(str).str.strip() != ""]
        work = work[~work["campaign"].astype(str).str.lower().str.contains("итого", na=False)]
    
    for col in ["spend", "clicks", "conversions", "impressions", "bounce_rate", "depth"]:
        if col in work.columns:
            work[col] = clean_numeric(work[col])
        else:
            work[col] = np.nan

    # "Отказы, %" обычно попадает как число типа 12.3 (без символа %).
    # В Excel формат '0.00%' ожидает долю (0.123), поэтому приводим.
    if "bounce_rate" in work.columns:
        s = work["bounce_rate"].dropna()
        if not s.empty and s.max() > 1.0:
            work["bounce_rate"] = work["bounce_rate"] / 100.0

    for col in ["campaign", "ad_group", "placement", "device", "gender", "age", "targeting_type", "title", "image", "can_block"]:
        if col not in work.columns:
            work[col] = "Не указано"
        work[col] = work[col].fillna("Не указано").astype(str).str.strip().replace({"": "Не указано"})

    work["spend"] = work["spend"].fillna(0.0)
    work["clicks"] = work["clicks"].fillna(0.0)
    work["conversions"] = work["conversions"].fillna(0.0)
    work["impressions"] = work["impressions"].fillna(0.0)

    return work


def determine_campaign_type(campaign_name: str) -> str:
    """Определяет тип кампании по её названию."""
    name = str(campaign_name)
    name_lower = name.lower()
    
    # Динамическая - наивысший приоритет
    if "динамическ" in name_lower:
        return "Динамическая"
    
    # Мастер кампаний - проверяем "МК " (с пробелом) в исходном регистре
    if "МК " in name:
        return "Мастер кампаний"
    
    # Остальные типы
    if "смарт" in name_lower:
        return "Смарт-баннеры"
    if "товарн" in name_lower:
        return "Товарная"
    if "рся" in name_lower:
        return "РСЯ"
    if "поиск" in name_lower:
        return "Поиск"
    
    return "Не определен"

def add_metrics(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["cpc"] = np.where(df["clicks"] > 0, df["spend"] / df["clicks"], np.nan)
    df["cpa"] = np.where(df["conversions"] > 0, df["spend"] / df["conversions"], np.nan)
    df["cr"] = np.where(df["clicks"] > 0, df["conversions"] / df["clicks"], np.nan)
    df["ctr"] = np.where(df["impressions"] > 0, df["clicks"] / df["impressions"], np.nan)
    # Добавляем тип кампании
    df["campaign_type"] = df["campaign"].apply(determine_campaign_type)
    return df


def summarize(df: pd.DataFrame, dim: str, by_campaign: bool = False) -> pd.DataFrame:
    """Агрегация данных по измерению, опционально с группировкой по кампаниям."""
    group_cols = ["campaign", dim] if by_campaign else [dim]
    
    # Для площадок добавляем информацию о возможности блокировки
    agg_dict = {
        "spend": ("spend", "sum"),
        "clicks": ("clicks", "sum"),
        "conversions": ("conversions", "sum"),
        "impressions": ("impressions", "sum"),
        "bounce_rate": ("bounce_rate", "mean"),
        "depth": ("depth", "mean"),
    }
    
    # Если это площадки и есть колонка can_block, добавляем её
    if dim == "placement" and "can_block" in df.columns:
        agg_dict["can_block"] = ("can_block", "first")
    
    agg = df.groupby(group_cols, dropna=False).agg(**agg_dict).reset_index()
    
    agg["cpc"] = np.where(agg["clicks"] > 0, agg["spend"] / agg["clicks"], np.nan)
    agg["cpa"] = np.where(agg["conversions"] > 0, agg["spend"] / agg["conversions"], np.nan)
    agg["cr"] = np.where(agg["clicks"] > 0, agg["conversions"] / agg["clicks"], np.nan)
    agg["ctr"] = np.where(agg["impressions"] > 0, agg["clicks"] / agg["impressions"], np.nan)
    
    # Сортировка: сначала по кампании (если есть), затем по конверсиям и расходу
    sort_cols = ["campaign", "spend"] if by_campaign else ["conversions", "spend"]
    agg = agg.sort_values(sort_cols, ascending=[True, False] if by_campaign else [False, False]).reset_index(drop=True)
    
    return agg

def summarize_gender_age(df: pd.DataFrame) -> pd.DataFrame:
    """Агрегация по полу и возрасту вместе."""
    if "gender" not in df.columns or "age" not in df.columns:
        return pd.DataFrame()
    
    # Создаем комбинированное поле
    df_copy = df.copy()
    df_copy["gender_age"] = df_copy["gender"].astype(str) + " " + df_copy["age"].astype(str)
    
    agg = df_copy.groupby(["campaign", "gender_age"], dropna=False).agg(
        spend=("spend", "sum"),
        clicks=("clicks", "sum"),
        conversions=("conversions", "sum"),
        impressions=("impressions", "sum"),
        bounce_rate=("bounce_rate", "mean"),
        depth=("depth", "mean"),
    ).reset_index()
    
    agg["cpc"] = np.where(agg["clicks"] > 0, agg["spend"] / agg["clicks"], np.nan)
    agg["cpa"] = np.where(agg["conversions"] > 0, agg["spend"] / agg["conversions"], np.nan)
    agg["cr"] = np.where(agg["clicks"] > 0, agg["conversions"] / agg["clicks"], np.nan)
    agg["ctr"] = np.where(agg["impressions"] > 0, agg["clicks"] / agg["impressions"], np.nan)
    
    agg = agg.sort_values(["campaign", "spend"], ascending=[True, False]).reset_index(drop=True)
    
    return agg


def overall_metrics(df: pd.DataFrame) -> Dict[str, float]:
    spend = df["spend"].sum()
    clicks = df["clicks"].sum()
    conv = df["conversions"].sum()
    impr = df["impressions"].sum()
    return {
        "spend": spend,
        "clicks": clicks,
        "conversions": conv,
        "impressions": impr,
        "cpc": spend / clicks if clicks else np.nan,
        "cpa": spend / conv if conv else np.nan,
        "cr": conv / clicks if clicks else np.nan,
        "ctr": clicks / impr if impr else np.nan,
    }


def fmt_money(x) -> str:
    if pd.isna(x):
        return "—"
    return f"{x:,.2f} руб".replace(",", " ")


def fmt_num(x) -> str:
    if pd.isna(x):
        return "—"
    return f"{int(round(x)):,}".replace(",", " ")


def fmt_pct(x) -> str:
    if pd.isna(x):
        return "—"
    return f"{x * 100:.2f}%"


def best_and_worst(summary: pd.DataFrame, min_clicks: int, min_spend: float) -> Tuple[pd.DataFrame, pd.DataFrame]:
    eligible = summary[(summary["clicks"] >= min_clicks) | (summary["spend"] >= min_spend)].copy()
    if eligible.empty:
        return eligible, eligible
    good = eligible[eligible["conversions"] > 0].sort_values(["cpa", "conversions"], ascending=[True, False]).head(5)
    bad = eligible[(eligible["conversions"] == 0) | (eligible["cpa"] > eligible["cpa"].median())].sort_values(["conversions", "spend"], ascending=[True, False]).head(5)
    return good, bad


def campaign_recommendations(
    df: pd.DataFrame,
    campaign_summary: pd.DataFrame,
    min_clicks: int,
    min_spend: float,
    available_dims: set,
) -> pd.DataFrame:
    account_cpa = overall_metrics(df)["cpa"]
    rows = []
    for _, row in campaign_summary.iterrows():
        camp = row["campaign"]
        part = df[df["campaign"] == camp].copy()
        recs = []

        for dim, label in [
            ("device", "устройствам"),
            ("gender", "полу"),
            ("age", "возрасту"),
            ("targeting_type", "типу условия показа"),
            ("placement", "площадкам"),
        ]:
            # Строим рекомендации только по тем измерениям, которые реально присутствуют в отчёте.
            if dim not in available_dims:
                continue

            s = summarize(part, dim)
            good, bad = best_and_worst(s, min_clicks=min_clicks, min_spend=min_spend)
            if not good.empty:
                good_names = ", ".join(good[dim].astype(str).head(3).tolist())
                recs.append(f"Усилить по {label}: {good_names}")
            if not bad.empty:
                bad_names = ", ".join(bad[dim].astype(str).head(3).tolist())
                recs.append(f"Сократить по {label}: {bad_names}")

        status = "усилить"
        if pd.notna(row["cpa"]) and pd.notna(account_cpa):
            if row["cpa"] > account_cpa * 1.5:
                status = "сократить и пересобрать"
            elif row["cpa"] > account_cpa * 1.15:
                status = "оптимизировать"

        rows.append({
            "campaign": camp,
            "spend": row["spend"],
            "clicks": row["clicks"],
            "conversions": row["conversions"],
            "cpc": row["cpc"],
            "cpa": row["cpa"],
            "cr": row["cr"],
            "status": status,
            "recommendations": " | ".join(recs[:10]) if recs else "Недостаточно данных для рекомендаций",
        })
    return pd.DataFrame(rows).sort_values(["conversions", "spend"], ascending=[False, False])


def print_terminal_report(df: pd.DataFrame, campaign_summary: pd.DataFrame, recs: pd.DataFrame) -> None:
    totals = overall_metrics(df)
    print("=" * 90)
    print("КРАТКИЙ ОТЧЕТ")
    print("=" * 90)
    print(f"Общий расход: {totals['spend']:.2f} руб")
    print(f"Клики: {int(totals['clicks'])}")
    print(f"Конверсии: {int(totals['conversions'])}")
    print(f"CPC: {totals['cpc']:.2f} руб")
    print(f"CPA: {totals['cpa']:.2f} руб")
    print(f"CR: {totals['cr']*100:.2f}%")
    print(f"\nВсего кампаний: {len(campaign_summary)}")
    print(f"Полный отчет сохранен в Excel и TXT файлах")
    print("=" * 90)


def safe_sheet_name(name: str) -> str:
    name = re.sub(r"[\\/*?:\[\]]", "_", str(name))
    return name[:31]


def write_df_sheet(ws, df: pd.DataFrame, table_name: str, sheet_title: str, report_title: str) -> None:
    """Записывает DataFrame на лист Excel с заголовками."""
    # Заголовок отчета в A1
    ws["A1"] = report_title
    ws["A1"].font = Font(size=18, bold=True)
    
    # Название вкладки в A2
    ws["A2"] = sheet_title
    ws["A2"].font = Font(size=16, bold=True)
    
    if df.empty:
        ws["A3"] = "Нет данных"
        return
    
    # Переименовываем колонку spend перед выводом
    df = df.copy()
    if "spend" in df.columns:
        df = df.rename(columns={"spend": "Расходы, руб с НДС"})
    
    start_row, start_col = 4, 1  # Начинаем с A4
    headers = list(df.columns)
    for i, h in enumerate(headers, start=start_col):
        c = ws.cell(row=start_row, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center")
    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, val in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=None if pd.isna(val) else val)

    end_row = start_row + len(df)
    end_col = start_col + len(headers) - 1
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    tab = Table(displayName=table_name[:25], ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    ws.freeze_panes = f"{get_column_letter(start_col)}{start_row + 1}"

    money_cols = {"Расходы, руб с НДС", "cpc", "cpa"}
    pct_cols = {"cr", "ctr", "bounce_rate"}
    int_cols = {"clicks", "conversions", "impressions"}

    for idx, h in enumerate(headers, start=start_col):
        col_letter = get_column_letter(idx)
        # Исправлена ошибка TypeError при вычислении ширины колонок
        max_len = df.iloc[:, idx - start_col].apply(lambda x: len(str(x)) if pd.notna(x) else 0).max() if len(df) else 0
        width = max(12, min(45, max(len(str(h)), max_len) + 2))
        ws.column_dimensions[col_letter].width = width
        if h in money_cols:
            for cell in ws[col_letter][start_row: end_row]:
                cell.number_format = '#,##0.00 [$₽-419]'
        elif h in pct_cols:
            for cell in ws[col_letter][start_row: end_row]:
                cell.number_format = '0.00%'
        elif h in int_cols:
            for cell in ws[col_letter][start_row: end_row]:
                cell.number_format = '#,##0'

    if "Расходы, руб с НДС" in headers:
        col_letter = get_column_letter(start_col + headers.index("Расходы, руб с НДС"))
        ws.conditional_formatting.add(
            f"{col_letter}{start_row+1}:{col_letter}{end_row}",
            ColorScaleRule(start_type='min', start_color='FFFFFF', end_type='max', end_color='BDD7EE')
        )
    if "cpa" in headers:
        col_letter = get_column_letter(start_col + headers.index("cpa"))
        ws.conditional_formatting.add(
            f"{col_letter}{start_row+1}:{col_letter}{end_row}",
            ColorScaleRule(start_type='min', start_color='63BE7B', mid_type='percentile', mid_value=50, mid_color='FFEB84', end_type='max', end_color='F8696B')
        )


def generate_recommendations_txt(output_path: Path, campaign_summary: pd.DataFrame, recs: pd.DataFrame, 
                                 placements_data: pd.DataFrame, report_title: str) -> None:
    """Генерирует TXT-файл с рекомендациями и списками площадок для минусовки."""
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f"{report_title}\n")
        f.write("=" * 80 + "\n\n")
        f.write("РЕКОМЕНДАЦИИ ПО УЛУЧШЕНИЮ КАМПАНИЙ\n")
        f.write("=" * 80 + "\n\n")
        
        for _, row in recs.iterrows():
            campaign = row['campaign']
            f.write(f"\nКАМПАНИЯ: {campaign}\n")
            f.write("-" * 80 + "\n")
            f.write(f"Статус: {row['status']}\n")
            f.write(f"Расход: {row['spend']:.2f} руб\n")
            f.write(f"Конверсии: {int(row['conversions'])}\n")
            f.write(f"CPA: {row['cpa']:.2f} руб\n\n")
            
            f.write("Рекомендации:\n")
            for rec in str(row['recommendations']).split(' | '):
                f.write(f"  • {rec}\n")
            
            # Площадки для минусовки (если есть данные)
            if not placements_data.empty:
                campaign_placements = placements_data[placements_data['campaign'] == campaign]
                bad_placements = campaign_placements[
                    (campaign_placements['conversions'] == 0) & 
                    (campaign_placements['spend'] > 100)
                ].copy()
                
                if not bad_placements.empty:
                    f.write("\nПЛОЩАДКИ ДЛЯ МИНУСОВКИ:\n")
                    for _, pl_row in bad_placements.iterrows():
                        placement_name = pl_row.get('placement', 'Неизвестная площадка')
                        can_block = pl_row.get('can_block', 'Да')
                        spend = pl_row.get('spend', 0)
                        
                        if str(can_block).lower() == 'нет':
                            f.write(f"  • {placement_name} (расход: {spend:.2f} руб) - ЭТУ ПЛОЩАДКУ ВЫКЛЮЧИТЬ НЕЛЬЗЯ\n")
                        else:
                            f.write(f"  • {placement_name} (расход: {spend:.2f} руб)\n")
            
            f.write("\n")
        
        f.write("\n" + "=" * 80 + "\n")
        f.write("Конец отчета\n")

def build_excel(output_path: Path, totals: Dict[str, float], campaign_summary: pd.DataFrame, recs: pd.DataFrame, 
                slices: Dict[str, pd.DataFrame], report_title: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Обзор"
    
    # Заголовки
    ws["A1"] = report_title
    ws["A1"].font = Font(size=18, bold=True)
    ws["A2"] = "Обзор"
    ws["A2"].font = Font(size=16, bold=True)
    
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18

    overview = [
        ("Общий расход, руб с НДС", totals["spend"]),
        ("Клики", totals["clicks"]),
        ("Конверсии", totals["conversions"]),
        ("Показы", totals["impressions"]),
        ("Средний CPC", totals["cpc"]),
        ("Средний CPA", totals["cpa"]),
        ("CR", totals["cr"]),
        ("CTR", totals["ctr"]),
    ]
    r = 5  # Начинаем после заголовков A1 и A2
    for name, value in overview:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=None if pd.isna(value) else float(value))
        r += 1
    for rr in range(5, r):
        label = ws.cell(row=rr, column=1).value
        cell = ws.cell(row=rr, column=2)
        if label in {"Общий расход, руб с НДС", "Средний CPC", "Средний CPA"}:
            cell.number_format = '#,##0.00 [$₽-419]'
        elif label in {"CR", "CTR"}:
            cell.number_format = '0.00%'
        else:
            cell.number_format = '#,##0'

    # Вкладка Кампании
    camp_sheet = wb.create_sheet("Кампании")
    write_df_sheet(camp_sheet, campaign_summary, "CampaignSummary", "Кампании", report_title)
    
    # Вкладка Тип кампании (вторая по порядку)
    if "campaign_type" in slices:
        type_sheet = wb.create_sheet("Тип кампании", 1)
        write_df_sheet(type_sheet, slices["campaign_type"], "CampaignType", "Тип кампании", report_title)
    
    # Вкладка Рекомендации
    rec_sheet = wb.create_sheet("Рекомендации")
    write_df_sheet(rec_sheet, recs, "CampaignRecs", "Рекомендации", report_title)

    # Остальные вкладки
    dimension_labels = {
        "device": "Устройство",
        "gender_age": "Пол и возраст",
        "targeting_type": "Тип условия показа",
        "placement": "Площадка",
        "ad_group": "Группа",
        "title": "Заголовок",
        "image": "Изображение",
    }
    
    for key, label in dimension_labels.items():
        if key in slices:
            sh = wb.create_sheet(safe_sheet_name(label))
            write_df_sheet(sh, slices[key], f"T_{key[:20]}", label, report_title)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    print("=" * 80)
    print("АНАЛИЗАТОР ОТЧЕТОВ ЯНДЕКС.ДИРЕКТ")
    print("=" * 80)
    print(f"\nНастройки:")
    print(f"  Сайт: {SITE_NAME}")
    print(f"  Заголовок отчета: {REPORT_TITLE}")
    print(f"  Папка с данными: {DATA_FOLDER}")
    print(f"  Папка для результатов: {OUTPUT_FOLDER}\n")
    
    try:
        # 1. Поиск файлов
        print("Шаг 1: Поиск файлов данных...")
        files = find_site_files(SITE_NAME, DATA_FOLDER)
        print()
        
        # 2. Загрузка основного файла (mainstat)
        print("Шаг 2: Загрузка основного файла mainstat...")
        required_cols = ["Название кампании", "Расход", "Клики", "Конверсии"]
        df_main = load_data_file(files["mainstat"], required_cols)
        
        # Маппинг колонок и подготовка данных
        mapping = map_columns(df_main)
        df = prepare_dataframe(df_main, mapping)
        df = add_metrics(df)
        print(f"  Обработано кампаний: {df['campaign'].nunique()}")
        print()
        
        # 3. Загрузка дополнительных файлов (опционально)
        additional_data = {}
        
        if files["regions"]:
            print(f"Шаг 3a: Загрузка данных по регионам...")
            try:
                df_regions = load_data_file(files["regions"], ["Название кампании", "Регион"])
                additional_data["regions"] = df_regions
            except Exception as e:
                print(f"  Ошибка загрузки regions: {e}")
        
        if files["polvozrast"]:
            print(f"Шаг 3b: Загрузка данных по полу и возрасту...")
            try:
                df_polvozrast = load_data_file(files["polvozrast"], ["Название кампании", "Пол", "Возраст"])
                additional_data["polvozrast"] = df_polvozrast
            except Exception as e:
                print(f"  Ошибка загрузки polvozrast: {e}")
        
        if files["pfrases"]:
            print(f"Шаг 3c: Загрузка данных по поисковым фразам...")
            try:
                df_pfrases = load_data_file(files["pfrases"], ["Название кампании", "Ключевая фраза"])
                additional_data["pfrases"] = df_pfrases
            except Exception as e:
                print(f"  Ошибка загрузки pfrases: {e}")
        print()
        
        # 4. Расчет общих метрик
        print("Шаг 4: Расчет метрик...")
        totals = overall_metrics(df)
        campaign_summary = summarize(df, "campaign")
        
        # available_dims: измерения из основного файла
        available_dims = set(mapping.keys())
        
        # 5. Создание срезов данных
        print("Шаг 5: Создание аналитических срезов...")
        slices = {}
        
        # Тип кампании
        slices["campaign_type"] = summarize(df, "campaign_type")
        print("  [OK] Тип кампании")
        
        # Устройство (по кампаниям)
        if "device" in available_dims:
            slices["device"] = summarize(df, "device", by_campaign=True)
            print("  [OK] Устройство (по кампаниям)")
        
        # Пол и возраст (объединенные)
        if "gender" in available_dims and "age" in available_dims:
            slices["gender_age"] = summarize_gender_age(df)
            print("  [OK] Пол и возраст (объединенные)")
        
        # Тип условия показа (по кампаниям)
        if "targeting_type" in available_dims:
            slices["targeting_type"] = summarize(df, "targeting_type", by_campaign=True)
            print("  [OK] Тип условия показа (по кампаниям)")
        
        # Площадка (по кампаниям)
        if "placement" in available_dims:
            slices["placement"] = summarize(df, "placement", by_campaign=True)
            print("  [OK] Площадка (по кампаниям)")
        
        # Группа (по кампаниям)
        if "ad_group" in available_dims:
            slices["ad_group"] = summarize(df, "ad_group", by_campaign=True)
            print("  [OK] Группа (по кампаниям)")
        
        # Заголовок (по кампаниям)
        if "title" in available_dims:
            slices["title"] = summarize(df, "title", by_campaign=True)
            print("  [OK] Заголовок (по кампаниям)")
        
        # Изображение (по кампаниям)
        if "image" in available_dims:
            slices["image"] = summarize(df, "image", by_campaign=True)
            print("  [OK] Изображение (по кампаниям)")
        print()
        
        # 6. Генерация рекомендаций
        print("Шаг 6: Формирование рекомендаций...")
        recs = campaign_recommendations(
            df,
            campaign_summary,
            MIN_CLICKS,
            MIN_SPEND,
            available_dims=available_dims,
        )
        print()
        
        # 7. Вывод отчета в терминал
        print_terminal_report(df, campaign_summary, recs)
        
        # 8. Сохранение Excel
        print("Шаг 7: Создание Excel-отчета...")
        output_excel = Path(OUTPUT_FOLDER) / f"{SITE_NAME}_analysis_result.xlsx"
        try:
            build_excel(output_excel, totals, campaign_summary, recs, slices, REPORT_TITLE)
            print(f"  [OK] Excel-отчет сохранен: {output_excel}")
        except PermissionError:
            print(f"  [ОШИБКА] Файл {output_excel} открыт в другой программе!")
            print(f"  Закройте файл и запустите скрипт снова.")
            raise
        
        # 9. Сохранение TXT с рекомендациями
        print("Шаг 8: Создание TXT-файла с рекомендациями...")
        output_txt = Path(OUTPUT_FOLDER) / f"{SITE_NAME}_recommendations.txt"
        placements_data = slices.get("placement", pd.DataFrame())
        generate_recommendations_txt(output_txt, campaign_summary, recs, placements_data, REPORT_TITLE)
        print(f"  [OK] TXT-файл сохранен: {output_txt}")
        
        print("\n" + "=" * 80)
        print("АНАЛИЗ ЗАВЕРШЕН УСПЕШНО!")
        print("=" * 80)
        
    except AnalyzerError as e:
        print(f"\n[ERROR] ОШИБКА: {e}")
        return 1
    except PermissionError as e:
        print(f"\n[ERROR] ФАЙЛ ЗАБЛОКИРОВАН")
        print(f"Закройте Excel-файл и повторите попытку.")
        return 1
    except Exception as e:
        print(f"\n[ERROR] НЕОЖИДАННАЯ ОШИБКА: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0


if __name__ == "__main__":
    main()
